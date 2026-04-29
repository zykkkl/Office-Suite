"""XLSX 渲染器 — IRDocument → .xlsx 文件

使用 openpyxl 将 IR 渲染为 Excel 工作簿。

架构位置：ir/compiler.py → IRDocument → [本文件] → .xlsx 文件

渲染流程：
  1. 创建 Workbook
  2. 遍历 IRDocument.children (SECTION 或 SLIDE)
  3. 每个节映射为一个 Sheet
  4. 元素按类型渲染到单元格/图表
  5. 保存 .xlsx 文件

XLSX 能力：
  - Sheet/数据写入
  - 图表（bar/column/line/pie/scatter）
  - 单元格合并
  - 数字格式
  - 条件格式
  - 冻结窗格
  - 样式（字体/填充/边框）
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

from ...ir.types import IRDocument, IRNode, IRStyle, NodeType
from ...ir.validator import validate_ir_v2
from ..base import BaseRenderer, RendererCapability

# openpyxl 图表类型映射
CHART_CLASS_MAP = {
    "bar": BarChart,
    "column": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "scatter": ScatterChart,
}

# 数字格式映射
NUMBER_FORMAT_MAP = {
    "currency": '#,##0.00 "¥"',
    "currency_usd": '"$"#,##0.00',
    "percent": "0.00%",
    "date": "YYYY-MM-DD",
    "datetime": "YYYY-MM-DD HH:MM:SS",
    "integer": "#,##0",
    "decimal_2": "#,##0.00",
    "decimal_4": "#,##0.0000",
    "scientific": "0.00E+00",
}


class XLSXRenderer(BaseRenderer):
    """Excel 工作簿渲染器"""

    def __init__(self):
        self._wb: Workbook | None = None
        self._current_row: int = 1

    @property
    def capability(self) -> RendererCapability:
        return RendererCapability(
            supported_node_types={
                NodeType.SLIDE, NodeType.SECTION, NodeType.TEXT,
                NodeType.TABLE, NodeType.CHART, NodeType.GROUP,
            },
            supported_layout_modes=set(),
            supported_text_transforms=set(),
            supported_animations=set(),
            supported_effects={"font_color", "fill_color", "bold", "border"},
            fallback_map={
                "gradient_fill": "solid_fill",
                "shadow": "none",
            },
        )

    def render(self, doc: IRDocument, output_path: str | Path) -> Path:
        """渲染 IRDocument 为 .xlsx 文件"""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        validation = validate_ir_v2(doc)
        for issue in validation.issues:
            print(f"[IR {issue.severity.value.upper()}] {issue}")

        self._wb = Workbook()
        # 删除默认 Sheet
        if "Sheet" in self._wb.sheetnames:
            del self._wb["Sheet"]

        for i, node in enumerate(doc.children):
            sheet_name = node.extra.get("title", f"Sheet{i+1}")
            ws = self._wb.create_sheet(title=sheet_name[:31])  # Sheet 名最长 31 字符
            self._current_row = 1
            self._render_node_to_sheet(node, ws, doc)

            # 冻结窗格
            freeze_row = node.extra.get("freeze_row")
            freeze_col = node.extra.get("freeze_col")
            if freeze_row or freeze_col:
                row_num = freeze_row or 1
                col_letter = get_column_letter(freeze_col) if freeze_col else "A"
                ws.freeze_panes = f"{col_letter}{row_num}"

        self._wb.save(str(output_path))
        return output_path

    def _render_node_to_sheet(self, node: IRNode, ws, doc: IRDocument):
        """将节点渲染到 Sheet"""
        if node.node_type == NodeType.SLIDE or node.node_type == NodeType.SECTION:
            for child in node.children:
                self._render_element_to_sheet(child, ws, doc)
        else:
            self._render_element_to_sheet(node, ws, doc)

    def _render_element_to_sheet(self, node: IRNode, ws, doc: IRDocument):
        """按节点类型渲染到 Sheet"""
        if node.node_type == NodeType.TEXT:
            self._render_text(node, ws)
        elif node.node_type == NodeType.TABLE:
            self._render_table(node, ws)
        elif node.node_type == NodeType.CHART:
            self._render_chart(node, ws)
        elif node.node_type == NodeType.GROUP:
            for child in node.children:
                self._render_element_to_sheet(child, ws, doc)
        else:
            ws.cell(row=self._current_row, column=1, value=f"[{node.node_type.value}]")
            self._current_row += 1

    def _render_text(self, node: IRNode, ws):
        """渲染文本 → 写入单元格

        支持 extra.column 指定列号。
        """
        col = node.extra.get("column", 1)
        cell = ws.cell(row=self._current_row, column=col, value=node.content or "")
        self._apply_cell_style(cell, node.style)
        # 数字格式
        number_format = node.extra.get("number_format", "")
        if number_format:
            cell.number_format = NUMBER_FORMAT_MAP.get(number_format, number_format)
        self._current_row += 1

    def _render_table(self, node: IRNode, ws):
        """渲染表格 → 写入数据区域

        支持：
        - 表头样式和斑马纹
        - 单元格合并（extra.merge_cells）
        - 数字格式（extra.number_format）
        - 条件格式（extra.conditional_format）
        """
        data = node.extra.get("data", [])
        rows = node.extra.get("rows", len(data))
        cols = node.extra.get("cols", len(data[0]) if data else 0)
        number_format = node.extra.get("number_format", "")
        merge_cells = node.extra.get("merge_cells", [])
        conditional_format = node.extra.get("conditional_format", "")

        start_row = self._current_row

        for r, row_data in enumerate(data[:rows]):
            if isinstance(row_data, list):
                for c, cell_val in enumerate(row_data[:cols]):
                    cell = ws.cell(
                        row=start_row + r,
                        column=c + 1,
                        value=cell_val,
                    )
                    # 表头样式
                    if r == 0:
                        cell.font = Font(bold=True, color="FFFFFF", size=11)
                        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                    elif r % 2 == 0:
                        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

                    cell.border = Border(
                        left=Side(style="thin", color="E2E8F0"),
                        right=Side(style="thin", color="E2E8F0"),
                        top=Side(style="thin", color="E2E8F0"),
                        bottom=Side(style="thin", color="E2E8F0"),
                    )

                    # 数字格式
                    if number_format and r > 0:  # 跳过表头
                        cell.number_format = NUMBER_FORMAT_MAP.get(number_format, number_format)

        # 单元格合并
        for merge_spec in merge_cells:
            # merge_spec: "A1:B2" 或 [start_row, start_col, end_row, end_col]
            if isinstance(merge_spec, str):
                ws.merge_cells(merge_spec)
            elif isinstance(merge_spec, list) and len(merge_spec) == 4:
                sr, sc, er, ec = merge_spec
                ws.merge_cells(
                    start_row=start_row + sr,
                    start_column=sc,
                    end_row=start_row + er,
                    end_column=ec,
                )

        # 条件格式
        if conditional_format and cols > 0:
            self._apply_conditional_format(ws, start_row, rows, cols, conditional_format)

        # 自动列宽
        for c in range(cols):
            col_letter = get_column_letter(c + 1)
            max_len = max(
                (len(str(row[c])) for row in data[:rows] if isinstance(row, list) and c < len(row)),
                default=8,
            )
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        self._current_row = start_row + rows + 1  # 空一行

    def _apply_conditional_format(self, ws, start_row: int, rows: int, cols: int, fmt_type: str):
        """应用条件格式

        fmt_type:
          - "data_bar": 数据条
          - "color_scale": 色阶（红-黄-绿）
          - "greater_than": 大于平均值高亮
        """
        # 条件格式应用到数据区域（跳过表头）
        if rows <= 1:
            return
        data_range = f"{get_column_letter(1)}{start_row + 1}:{get_column_letter(cols)}{start_row + rows - 1}"

        if fmt_type == "data_bar":
            rule = DataBarRule(
                start_type="min",
                end_type="max",
                color="638EC6",
            )
            ws.conditional_formatting.add(data_range, rule)
        elif fmt_type == "color_scale":
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            )
            ws.conditional_formatting.add(data_range, rule)
        elif fmt_type == "greater_than":
            rule = CellIsRule(
                operator="greaterThan",
                formula=[f"AVERAGE({data_range})"],
                fill=PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            )
            ws.conditional_formatting.add(data_range, rule)

    def _render_chart(self, node: IRNode, ws):
        """渲染图表 → 在 Sheet 中嵌入图表

        支持图表类型：bar, column, line, pie, scatter
        支持数据标签、图例位置。
        """
        chart_type_str = node.chart_type or node.extra.get("chart_type", "bar")
        categories = node.extra.get("categories", [])
        series_list = node.extra.get("series", [])
        title = node.extra.get("title", "")
        show_legend = node.extra.get("show_legend", True)
        show_data_labels = node.extra.get("show_data_labels", False)

        if not series_list:
            ws.cell(row=self._current_row, column=1, value="[图表: 无数据]")
            self._current_row += 1
            return

        # 写入数据到隐藏区域
        data_start_row = self._current_row
        # 表头
        ws.cell(row=data_start_row, column=1, value="类别")
        for i, s in enumerate(series_list):
            ws.cell(row=data_start_row, column=i + 2, value=s.get("name", f"系列{i+1}"))
        # 数据
        for r, cat in enumerate(categories):
            ws.cell(row=data_start_row + 1 + r, column=1, value=cat)
            for i, s in enumerate(series_list):
                values = s.get("values", [])
                if r < len(values):
                    ws.cell(row=data_start_row + 1 + r, column=i + 2, value=values[r])

        data_rows = len(categories)

        # 创建图表
        chart_class = CHART_CLASS_MAP.get(chart_type_str, BarChart)
        chart = chart_class()
        chart.title = title
        chart.style = 10
        # PieChart 没有 axis 属性
        if not isinstance(chart, (PieChart, ScatterChart)):
            chart.y_axis.title = ""
            chart.x_axis.title = ""

        # 数据引用
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + data_rows)
        for i in range(len(series_list)):
            data_ref = Reference(
                ws,
                min_col=i + 2,
                min_row=data_start_row,
                max_row=data_start_row + data_rows,
            )
            chart.add_data(data_ref, titles_from_data=True)

        # ScatterChart 使用 xvalues 而非 categories
        if chart_type_str == "scatter":
            # scatter: categories 作为 X，每个系列作为 Y
            if chart.series:
                x_ref = Reference(ws, min_col=1, min_row=data_start_row + 1, max_row=data_start_row + data_rows)
                for s in chart.series:
                    s.xvalues = x_ref
        else:
            chart.set_categories(cats_ref)

        # 数据标签
        if show_data_labels:
            for s in chart.series:
                s.dLbls = DataLabelList()
                s.dLbls.showVal = True

        # 图例
        if not show_legend:
            chart.legend = None

        # 图表尺寸
        chart.width = 20
        chart.height = 12

        # 嵌入图表
        ws.add_chart(chart, f"A{self._current_row + data_rows + 2}")
        self._current_row += data_rows + 18  # 图表占位

    def _apply_cell_style(self, cell, style: IRStyle | None):
        """应用单元格样式

        支持：字体大小、粗体、斜体、颜色、填充色、对齐
        """
        if style is None:
            return
        font_kwargs = {}
        if style.font_size:
            font_kwargs["size"] = style.font_size
        if style.font_weight and style.font_weight >= 700:
            font_kwargs["bold"] = True
        if style.font_italic:
            font_kwargs["italic"] = True
        if style.font_color:
            font_kwargs["color"] = style.font_color.lstrip("#")
        if font_kwargs:
            cell.font = Font(**font_kwargs)

        # 填充色
        if style.fill_color:
            cell.fill = PatternFill(
                start_color=style.fill_color.lstrip("#"),
                end_color=style.fill_color.lstrip("#"),
                fill_type="solid",
            )

        # 边框
        if style.border:
            border_color = style.border.get("color", "000000").lstrip("#")
            border_width = style.border.get("width", 1)
            side_style = "medium" if border_width >= 2 else "thin"
            cell.border = Border(
                left=Side(style=side_style, color=border_color),
                right=Side(style=side_style, color=border_color),
                top=Side(style=side_style, color=border_color),
                bottom=Side(style=side_style, color=border_color),
            )
